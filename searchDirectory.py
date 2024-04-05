import os
import re
from openpyxl import Workbook

# This script doesn't separate your ascx files and c# files and also 
# shows when the folder changes on the excel sheet

entries = [
  (
    os.path.abspath("../python/String-Finder-1.0.0/Page Template Library"),
    ["Page Template Library: .ascx and c#", 
    [r'SqlDataAdapter\("', r'SqlCommand\("', r'"UPDATE ', r'"update ', r'"SELECT ', r'"select ', r'"DELETE ', r'"delete ', r'"INSERT ', r'"insert ', r"<asp:SqlDataSource"],
    "page-template-library.xlsx"],
  )
  ]
    
def save_to_wb(directory: str, excel_props: tuple) -> None:
    sql_src = handle_all_files(directory, excel_props[1])
    # Don't create excel sheet if no patterns match 
    if len(sql_src) == 0:
        print(os.listdir(directory))
        return
        # If needed we can swap return for an Exception but it can be problematic
        raise Exception("Your excel sheet will be null. Terminating...")
    wb = Workbook()
    ws = wb.active
    ws.append(["Title: ", excel_props[0]])
    ws.append(["Pattern: ", str(excel_props[1])])
    ws.append(["SQL Calls: ", len(sql_src)])
    currentDir = ""
    currentFile = ""
    fileCount = 0
    uniqueNames = []
    for root, file, line_number, command, pattern in sql_src:
        # If you are entering a new folder path, show that in the excel sheet
        if (root != currentDir):
            currentDir = root
            ws.append([""])
            ws.append([""])
            ws.append(["PATH:", root])
            ws.append(["File", "Line Number","Pattern Matched", "Select Command"])
        if (file != currentFile ):
            uniqueNames.append(file)
            currentFile = file
            fileCount +=1
        ws.append([file, line_number, pattern, command])

    ws["A4"] = "Unique Files: "
    ws["B4"] = fileCount
    wb.save(excel_props[2])
    print(f"Excel file '{excel_props[2]}' has been created.")
    return

# If we want to exclude SqlCommand("") that aren't 
# a stored procedure we can update the pattern
def handle_all_files(directory, patterns):
    # Searches all folders for files in your directory
    sql_to_excel = []
    for root, folder, files in os.walk(directory):
        for file in files:
            if not file.endswith(".designer.cs"):
                handle_os_walk(patterns, root, file, sql_to_excel)      
    return sql_to_excel

def handle_os_walk(patterns, root, file, sql_to_excel):
    # Searches your files for the specified SQL call pattern
    filepath = os.path.join(root, file)
    # you will get charmap codec can't decode byte XXXX errors if encoding not set properly
    with open(filepath, encoding="utf8") as f:
        content = f.readlines()
        for line_number, line in enumerate(content, 1):
            for pattern in patterns:
                # Trying to exclude commented out lines (lines that start with /)
                if re.match(r'^/', line.strip()): 
                    continue
                if re.search(pattern, line):
                    sql_to_excel.append(
                        (root, file, line_number, line.strip(), pattern)
                    )
                    # Break to prevent duplicate entries if diff patterns observed on same line
                    break
    return sql_to_excel

# Start script and spread entry properties in save_to_wb function
for entry in entries:
    save_to_wb(*entry)
